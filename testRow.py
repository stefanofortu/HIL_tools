from openpyxl import load_workbook
import xlsxwriter


def isStringCANSignal(stringValue):
    if "[" in stringValue and "]" in stringValue:
        return True
    else:
        return False


def formatFiles(tcOutputFilename):
    wb = load_workbook(tcOutputFilename)
    ws = wb["TC"]
    # Create a book with xlsxwriter
    optbook = xlsxwriter.Workbook(tcOutputFilename)
    # Add sheet with xlsxwriter
    optsheet = optbook.add_worksheet("TC")
    # Define format

    red = optbook.add_format({'color': 'red'})
    gray = optbook.add_format({'color': 'gray'})
    lightgray = optbook.add_format({'color': '#CCCCCC'})
    graySmall = optbook.add_format({'color': '#CCCCCC', 'font_size': 8})
    black = optbook.add_format({'color': 'black', 'font_size': 12, 'bold': True})

    for col in ws.iter_cols(min_row=2, max_row=23, min_col=2, max_col=2, values_only=True):
        for rowNum, cellString in enumerate(col):
            tmp_array = []
            # print(type(cellString))
            if isinstance(cellString, str):
                # print(cellString)
                splitvalue = cellString.split('\n')
                for row in splitvalue:
                    if row != "":
                        if isStringCANSignal(row):
                            tmp_array.append(graySmall)
                        else:
                            tmp_array.append(black)
                        tmp_array.append(row + "\n")
                # print(tmp_array)
                # Write in rich text
                # optsheet.write_rich_string('A1', red, splitvalue[0], splitvalue[1])
                optsheet.write_rich_string(rowNum + 1, 2, *tmp_array)
            # Split characters
            # print(cell.value)
    optbook.close()


def createTestOutput(tcInputFilename, tcOutputFilename, modification=True):
    wb = load_workbook(tcInputFilename)
    ws = wb["TC"]
    stringTC = ""
    for i in range(4):
        stringTC += ("hello" + str(i) + "\n")
    if modification:
        ws['A1'] = stringTC
    wb.save(filename=tcOutputFilename)


if __name__ == '__main__':
    createTestOutput(tcInputFilename='test.xlsx', tcOutputFilename='testOUT.xlsx', modification=True)
    formatFiles('testOUT.xlsx')
    createTestOutput(tcInputFilename='testOUT.xlsx', tcOutputFilename='testOUT2.xlsx', modification=False)

    exit()
    # cellvalue = stringTC
    # Split characters
    splitvalue = cellvalue.split('\n')
    tmp_array = []
    for row in splitvalue:
        if row != "":
            if '1' in row:
                tmp_array.append(red)
            tmp_array.append(row + "\n")

    print(tmp_array)
    # Write in rich text
    # optsheet.write_rich_string('A1', red, splitvalue[0], splitvalue[1])
    optsheet.write_rich_string('A1', *tmp_array)

    # End
    optbook.close()
