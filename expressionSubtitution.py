from utils.fileTemplateConfiguration import file_TC_BUILD_Column
from utils.excelUtils import getColumnIndexFromString, getColumnLetterFromString
from importDictionary import singleTestStep
from copy import copy
from openpyxl.utils import get_column_letter

def findExpressions(worksheetStart, substitutionDictionary):
    for rowNumStart, row in enumerate(worksheetStart.iter_rows(values_only=True)):
        for colNum, cell in enumerate(row):
            if isinstance(cell, str) and len(cell) > 2:
                if "()" in cell:
                    if cell not in substitutionDictionary:
                        print("function " + cell + " not found in dictionary")


def substituteFunctions(worksheetStart, worksheetEnd, substitutionDictionary):
    findExpressions(worksheetStart=worksheetStart,substitutionDictionary=substitutionDictionary)

    columnEnableIndex = getColumnIndexFromString(worksheetStart, file_TC_BUILD_Column['enable_header'])
    columnTestNIndex = getColumnIndexFromString(worksheetStart, file_TC_BUILD_Column['testN_header'])
    columnTestIDIndex = getColumnIndexFromString(worksheetStart, file_TC_BUILD_Column['testID_header'])
    columnDescriptionIndex = getColumnIndexFromString(worksheetStart, file_TC_BUILD_Column['stepDescr_header'])
    columnPreconditionIndex = getColumnIndexFromString(worksheetStart, file_TC_BUILD_Column['precondition_header'])
    columnActionIndex = getColumnIndexFromString(worksheetStart, file_TC_BUILD_Column['action_header'])
    columnExpectedResIndex = getColumnIndexFromString(worksheetStart, file_TC_BUILD_Column['expected_header'])

    columnDescriptionLetter = getColumnLetterFromString(worksheetStart, file_TC_BUILD_Column['stepDescr_header'])
    columnPreconditionLetter = getColumnLetterFromString(worksheetStart, file_TC_BUILD_Column['precondition_header'])
    columnActionLetter = getColumnLetterFromString(worksheetStart, file_TC_BUILD_Column['action_header'])
    columnExpectedResLetter = getColumnLetterFromString(worksheetStart, file_TC_BUILD_Column['expected_header'])

    # startColIndex = column_index_from_string(startCol)
    # endColIndex = column_index_from_string(endCol)
    c = 0
    translation = worksheetStart.max_row + 15
    rowNumEnd = translation
    rowsAdded = 0
    for rowNumStart, row in enumerate(worksheetStart.iter_rows(), start=1):
        # print(c)
        # c += 1
        foundStringToSubstitute = False
        for colNum, cell in enumerate(row, start=1):
            if isinstance(cell.value, str):
                # print(wsheetEnd.cell(row=rowNumEnd, column=colNum + startColIndex).value)
                if len(cell.value) > 2:
                    if "()" in cell.value:
                        functionParsed = parseFunction(cell.value)
                        functionName = functionParsed['functionName']
                        parametersList = functionParsed['parametersList']
                        substitutedFunction = createFunction(functionParsed, substitutionDictionary)
                        if functionName in substitutionDictionary:
                            foundStringToSubstitute = True
                            columnEnableValue = worksheetStart.cell(row=rowNumStart, column=columnEnableIndex).value
                            columnTestNValue = worksheetStart.cell(row=rowNumStart, column=columnTestNIndex).value
                            columnTestIDValue = worksheetStart.cell(row=rowNumStart, column=columnTestIDIndex).value
                            # parametersList
                            #for dataNum, t in enumerate(substitutionDictionary[functionName]['data']):
                            for dataNum, t in enumerate(substitutedFunction):
                                # riporto il valore di "step" in
                                newRowPos = rowNumStart + translation + rowsAdded + dataNum
                                # riporto il valore in testEnable
                                worksheetEnd.cell(row=newRowPos, column=columnEnableIndex, value=columnEnableValue)
                                # riporto il valore in testN
                                worksheetEnd.cell(row=newRowPos, column=columnTestNIndex, value=columnTestNValue)
                                # riporto il valore in testID
                                worksheetEnd.cell(row=newRowPos, column=columnTestIDIndex, value=columnTestIDValue)

                                # riporto il valore della descrizione nella colonna della step description
                                worksheetEnd.cell(row=newRowPos, column=columnDescriptionIndex, value=t.descr)
                                # riporto il valore dell'azione/risultato atteso" nella colonna corrente
                                worksheetEnd.cell(row=newRowPos, column=colNum, value=t.act)
                                # riporto il valore di "expected res" nella colonna degli expected results
                                worksheetEnd.cell(row=newRowPos, column=columnExpectedResIndex, value=t.exp)

                                if cell.has_style:
                                    worksheetEnd[columnDescriptionLetter + str(newRowPos)].fill = copy(cell.fill)
                                    worksheetEnd[columnDescriptionLetter + str(newRowPos)].border = copy(cell.border)
                                    worksheetEnd[columnDescriptionLetter + str(newRowPos)].font = copy(cell.font)
                                    worksheetEnd[columnPreconditionLetter + str(newRowPos)].fill = copy(cell.fill)
                                    worksheetEnd[columnPreconditionLetter + str(newRowPos)].border = copy(cell.border)
                                    worksheetEnd[columnPreconditionLetter + str(newRowPos)].font = copy(cell.font)
                                    worksheetEnd[columnActionLetter + str(newRowPos)].fill = copy(cell.fill)
                                    worksheetEnd[columnActionLetter + str(newRowPos)].border = copy(cell.border)
                                    worksheetEnd[columnActionLetter + str(newRowPos)].font = copy(cell.font)
                                    worksheetEnd[columnExpectedResLetter + str(newRowPos)].fill = copy(cell.fill)
                                    worksheetEnd[columnExpectedResLetter + str(newRowPos)].border = copy(cell.border)
                                    worksheetEnd[columnExpectedResLetter + str(newRowPos)].font = copy(cell.font)

                            #rowsAdded += len(substitutionDictionary[functionName]['data']) - 1
                            rowsAdded += len(substitutedFunction) - 1
        if not foundStringToSubstitute:
            worksheetEnd.move_range("A" + str(rowNumStart) + ":AA" + str(rowNumStart),
                                    rows=translation + rowsAdded, cols=0)

    worksheetEnd.delete_rows(1, amount=translation)


def parseFunction(stringToParse):
    if "=" in stringToParse:
        [functionName, parameters] = stringToParse.split("=")
        if ";" in parameters:
            parametersList = parameters.split(";")
        else:
            parametersList = [parameters]
    else:
        functionName = stringToParse
        parametersList = []

    return {"functionName": functionName, 'parametersList': parametersList}


def createFunction(functionParsed, substitutionDictionary):
    functionParsedName = functionParsed['functionName']
    parametersParsedList = functionParsed['parametersList']
    substitutedFunction = []
    if functionParsedName in substitutionDictionary:
        if len(substitutionDictionary[functionParsedName]['parameters']) == len(parametersParsedList):
            for functionStep in substitutionDictionary[functionParsedName]['data']:
                descr = functionStep.descr
                act = functionStep.act
                exp = functionStep.exp
                for parIndex, par in enumerate(substitutionDictionary[functionParsedName]['parameters']):
                    if isinstance(descr, str):
                        descr = descr.replace("{" + par + "}", parametersParsedList[parIndex])
                    if isinstance(act, str):
                        act = act.replace("{" + par + "}", parametersParsedList[parIndex])
                    if isinstance(exp, str):
                        exp = exp.replace("{" + par + "}", parametersParsedList[parIndex])
                s = singleTestStep(descr, act, exp)
                substitutedFunction.append(s)
            print(substitutedFunction)
        else:
            print(functionParsedName)
            raise ValueError
    else:
        print(functionParsedName + "not found")
    print("FIX THIS PART")
        # raise KeyError
    return substitutedFunction