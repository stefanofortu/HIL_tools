from openpyxl import Workbook
from importDictionary import importDictionary
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from expressionSubtitution import substituteExpressionsInCol, substituteExpressionsByRows

source_filename = 'Ferrari_TC_AF.xlsx'
destination_filename = 'Ferrari_TC_AF_v2.xlsx'

wbStart = load_workbook(source_filename)
wbStart.save(filename=destination_filename)
wbEnd = load_workbook(destination_filename)
wsStart = wbStart['TestCaseAF']
wsEnd = wbEnd['TestCaseAF']

import re

# Split the string at every white-space character:

txt = "NBC::[C1].STATUS_DOORS.STS_DRS_PsngrLatchsts=0x1(\"Unlocked\")"
# .*::\[.*\]\..*\..*=0x.\(\".*\"\)


for rowNum, row in enumerate(wsEnd.iter_rows(values_only=True), start=1):
    for colNum, cell in enumerate(row, start=1):
        if type(cell) == str:
            if "::" in cell:
                result = re.match('.*::\[.*\]\..*\..*=0x.\(\".*\"\)', cell)
                if result:
                    canValues = re.split('::|\[|\]|\.|=|\(|\)|\"', cell)
                    node = canValues[0]
                    can = canValues[2]
                    msg = canValues[4]
                    signal = canValues[5]
                    valueHex = canValues[6]
                    valueStr = canValues[8]
                    # print(msg)
                    newCellValue = "CAN[" + node + "." + can + "." + msg + "." + signal + "] = " + valueHex + " // " + valueStr
                    # cell = "hello"
                    try:
                        wsEnd.cell(column=colNum, row=rowNum, value=newCellValue)  # "{0}".format(get_column_letter(col)))
                    except:
                        print(cell)
                    # print("CAN[" + node + "." + can + "." + msg + "." + signal + "]=" + valueHex + " " + valueStr)
                # CAN [BRAKE_FE3.VehicleSpeedVSOSig] = Vnos

wbEnd.save(filename=destination_filename)
