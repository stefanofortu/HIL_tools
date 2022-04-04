from fillers import fillTestNColumn, fillEnableColumn, fillStepIDCounter
from importDictionary import importDictionaryV2
from utils.expressionSubtitution import substituteFunctions, removeTestTypeColumn, findExpressions
from utils.fileImporter import importFunctionFiles, importBuildFile, generateRunFileFromBuildFile
import sys
import json
from openpyxl import load_workbook
from itertools import groupby


def replaceCell(stringList, find_array, replace_array):
    outCell = []
    workCell = []
    for elem in stringList:
        workCell.append(elem)

    row_num = 0
    while row_num < len(workCell):
        if row_num < (len(workCell) - len(find_array)):
            match_counter = 0
            for find_array_index in range(0, len(find_array)):
                if workCell[row_num + find_array_index] == find_array[find_array_index]:
                    match_counter += 1
            if match_counter == len(find_array):
                # print("====")
                # print(workCell[row_num-1])
                # print(workCell[row_num])
                # print(workCell[row_num+1])
                # print("====")
                # print(len(find_array))
                for replace_array_index in range(0, len(replace_array)):
                    outCell.append(replace_array[replace_array_index])
                row_num += len(find_array)
            else:
                outCell.append(workCell[row_num])
                row_num += 1
        else:
            outCell.append(workCell[row_num])
            row_num += 1

    workCell = []
    for elem in outCell:
        workCell.append(elem)

    return outCell


def execCANinsertion():
    try:
        with open('pathFile.json', 'r') as json_file:
            json_file_no_comment = ''.join(line for line in json_file if not line.startswith('#'))
            json_data = json.loads(json_file_no_comment)

    except FileNotFoundError:
        print('File pathFile.json does not exist')
        sys.exit()

    find_replace_json = json_data['root']['find_replace_multiple_row']

    # rootPathFile = json.load(find_replace_json)
    input_file = find_replace_json['input_file']
    input_file_path = input_file['path']
    print('filePath for input file :', input_file_path)
    input_file_sheet = input_file['sheet_name']
    print('sheets in input file :', input_file_sheet)

    find_replace_file = find_replace_json['find_replace_file']
    find_replace_file_path = find_replace_file['path']
    print('filePath for find&Replace file :', find_replace_file_path)
    find_replace_file_sheet = find_replace_file['sheet_name']
    print('sheets in find&Replace file :', find_replace_file_sheet)

    output_file = find_replace_json['output_file']
    output_file_path = output_file['path']
    print('filePath for output file :', output_file_path)

    # print('findArray for output file :', findArray)
    # print('replaceArray for output file :', replaceArray)
    # Closing json file

    wb_in = load_workbook(input_file_path)
    ws_in = wb_in[input_file_sheet]

    print("import input file : DONE")

    wb_find_replace = load_workbook(find_replace_file_path)
    ws_find_replace = wb_find_replace[find_replace_file_sheet]

    print("open find replace file : DONE")
    substitution_list_from_excel = []
    for row in ws_find_replace.iter_rows(min_row=2, min_col=1, max_col=2):
        single_substitution = {"find": "string", "replace": "string"}
        for colNum, cell in enumerate(row):
            if isinstance(cell.value, str):
                if colNum == 0:
                    single_substitution["find"] = cell.value.split("\n")
                elif colNum == 1:
                    single_substitution["replace"] = cell.value.split("\n")
        substitution_list_from_excel.append(single_substitution)

    print(substitution_list_from_excel)
    print("import find replace file : DONE")

    wb_in.save(filename=output_file_path)
    wbOut = load_workbook(output_file_path)
    wsOut = wbOut[input_file_sheet]

    print("saved Copy of Input file: DONE")

    for substitution in substitution_list_from_excel:
        find_array = substitution['find']
        replace_array = substitution['replace']

        for col in wsOut.iter_cols(min_row=1, max_row=100, min_col=18, max_col=20):
            for rowNum, cell in enumerate(col):
                # tmp_array = []
                # print(type(cellString))
                if isinstance(cell.value, str):
                    cellString = cell.value
                    # print(len(cellString))
                    splitvalue = cellString.split('\n')
                    newCellValue = replaceCell(splitvalue, find_array, replace_array)
                    cellNewValueString = '\n'.join(newCellValue)
                    # print(len(cellNewValueString))
                    # print(cellNewValueString)
                    cell.value = str(cellNewValueString)
                    # print(cellNewValueString)
                    # tmp_array.append(row + "\n")
                    # print(tmp_array)
                    # Write in rich text
                    # optsheet.write_rich_string('A1', red, splitvalue[0], splitvalue[1])
                    # for elem in newCellValue:
                    # print("==================")
                    # optsheet.write_rich_string(rowNum + 1, 2, *tmp_array)
                # Split characters
                # print(cell.value)
        # optbook.close()

    print("substitution : DONE")

    for col in wsOut.iter_cols(min_row=1, max_row=1640, min_col=18, max_col=20):
        for cell in col:
            if isinstance(cell.value, str):
                cellString = cell.value
                splitvalue = cellString.split('\n')
                rowToRemove = []
                # for i, elem in enumerate(splitvalue):
                #    print(str(i) + " : " + elem)
                # print("len :" + str(len(splitvalue)))
                for i in range(0, len(splitvalue)):
                    if splitvalue[i] == "":
                        if i < len(splitvalue) - 1:
                            if splitvalue[i + 1] == "":
                                rowToRemove.append(i)
                        else:
                            rowToRemove.append(i)
                rowToRemove.reverse()
                # print(rowToRemove)
                for r in rowToRemove:
                    splitvalue.pop(r)
                # print(splitvalue)

                # newCellValue = groupby(splitvalue)
                # print(newCellValue)
                # newCellValue = replaceCell(splitvalue, find_array, replace_array)
                cellNewValueString = '\n'.join(splitvalue)
                # print(len(cellNewValueString))
                # print(cellNewValueString)

                cell.value = str(cellNewValueString)
    print("cancellazioni righe vuote : DONE")

    wbOut.save(filename=output_file_path)

    print("file output saving : DONE")


# exit()

# disableSequences(worksheet=wsRun)
# removeTestTypeColumn(worksheet=wsRun)
# fillTestNColumn(worksheet=wsRun)
# fillEnableColumn(worksheet=wsRun)
# fillStepIDCounter(worksheet=wsRun)

# print("other operations")

# " Salva"
# wbRun.save(filename=run_filename)

# print("file saved")

# sys.exit(0)


if __name__ == '__main__':
    execCANinsertion()
