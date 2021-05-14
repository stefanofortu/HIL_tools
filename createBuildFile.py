from openpyxl import Workbook
from importDictionary import importDictionary
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


source_filename = "C:\\Users\\Stefano\\Desktop\\WIP\\HIL\\TC_AF\\2-TC_Build\\TC_AF_HIL_Build_2021_05_13.xlsx"
destination_filename = "C:\\Users\\Stefano\\Desktop\\WIP\\HIL\\TC_AF\\3-TC_Run\\TC_AF_HIL_Run_2021_05_13.xlsx"

wbTCManual = load_workbook(source_filename)
#wbStart.save(filename=destination_filename)
wbTCBuild = load_workbook(destination_filename)

wsTCManual = wbTCManual['Template']
wsTCBuild = wbTCBuild['TC_HIL']

for colNum, cell in enumerate(wsTCManual[1], start=1):
    if cell.value == "PRECONDITIONS":
        print("precondition in colonna ", colNum, " lettera ", get_column_letter(colNum))
        pass
    elif cell.value == "ACTIONS":
        pass
    elif cell.value == "EXPECTED RESULT":
        pass

#wsStart = wbStart['test']
#wsEnd = wbEnd['test']

exit()