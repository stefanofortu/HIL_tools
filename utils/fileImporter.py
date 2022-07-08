from utils.verifyTemplates import checkActionFile, checkBuildFile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelFile:
    def __init__(self, tc_file_name, tc_sheet_name):
        self.file_name = tc_file_name
        self.sheet_name = tc_sheet_name
        self.workbook = load_workbook(self.file_name)
        self.worksheet = self.workbook[self.sheet_name]

    def getColumnIndexFromString(self, columnName):
        for row in self.worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for colNum, cell in enumerate(row, start=1):
                if cell == columnName:
                    return colNum
        print("Column " + columnName + " not found")
        raise ValueError

    def getColumnLetterFromString(self, columnName):
        for row in self.worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for colNum, cell in enumerate(row, start=1):
                if cell == columnName:
                    return get_column_letter(colNum)
        print("Column " + columnName + " not found")
        raise ValueError

    def save(self):
        self.workbook.save(filename=self.file_name)

    def save_copy(self, file_name):
        self.workbook.save(filename=file_name)

class TC_Build_File(ExcelFile):
    def __init__(self, tc_file_name, tc_sheet_name):
        super().__init__(tc_file_name, tc_sheet_name)
        try:
            self.enable_col_letter = self.getColumnLetterFromString("ENABLE")
            self.testN_col_letter = self.getColumnLetterFromString("TEST N.")
            self.test_ID_col_letter = self.getColumnLetterFromString("TEST ID")
            self.test_description_col_letter = self.getColumnLetterFromString("STEP DESCRIPTION")
            # self.tc_type_col_letter = self.getColumnLetterFromString("TC Type")
            self.ID_col_letter = self.getColumnLetterFromString("ID")
            self.precondition_col_letter = self.getColumnLetterFromString("PRECONDITIONS / OPERATION")
            self.actions_col_letter = self.getColumnLetterFromString("ACTIONS")
            self.expected_result_col_letter = self.getColumnLetterFromString("EXPECTED RESULT")
            self.time_step_col_letter = self.getColumnLetterFromString("Time step [ms]")
            self.sample_time_col_letter = self.getColumnLetterFromString("Sample time [ms]")
            self.tolerance_col_letter = self.getColumnLetterFromString("Tolerance [%]")

        except ValueError:
            print("In file --->", self.file_name)
            exit()



class TC_Run_File(ExcelFile):
    def __init__(self, tc_file_name, tc_sheet_name):
        super().__init__(tc_file_name, tc_sheet_name)
        try:
            self.enable_col_letter = self.getColumnLetterFromString("ENABLE")
            self.testN_col_letter = self.getColumnLetterFromString("TEST N.")
            self.test_ID_col_letter = self.getColumnLetterFromString("TEST ID")
            self.test_description_col_letter = self.getColumnLetterFromString("STEP DESCRIPTION")
            #self.tc_type_col_letter = self.getColumnLetterFromString("TC Type")
            self.ID_col_letter = self.getColumnLetterFromString("ID")
            self.precondition_col_letter = self.getColumnLetterFromString("PRECONDITIONS / OPERATION")
            self.actions_col_letter = self.getColumnLetterFromString("ACTIONS")
            self.expected_result_col_letter = self.getColumnLetterFromString("EXPECTED RESULT")
            self.time_step_col_letter = self.getColumnLetterFromString("Time step [ms]")
            self.sample_time_col_letter = self.getColumnLetterFromString("Sample time [ms]")
            self.tolerance_col_letter = self.getColumnLetterFromString("Tolerance [%]")
        except ValueError:
            print("In file ", self.file_name)
            exit()




class Functions_File(ExcelFile):
    def __init__(self, tc_file_name, tc_sheet_name):
        super().__init__(tc_file_name, tc_sheet_name)
        try:
            self.fieldType_col_letter = self.getColumnIndexFromString('fieldType')
            self.fieldType_col_index = self.getColumnIndexFromString('fieldType')
            self.value_col_letter = self.getColumnLetterFromString("value")
            self.enable_col_letter = self.getColumnLetterFromString("ENABLE")
            self.step_description_col_letter = self.getColumnLetterFromString("STEP DESCRIPTION")
            self.precondition_col_letter = self.getColumnLetterFromString("PRECONDITIONS/ACTION")
            self.expected_result_col_letter = self.getColumnLetterFromString("EXPECTED RESULTS")
            self.time_step_col_letter = self.getColumnLetterFromString("Time step [ms]")
            self.sample_time_col_letter = self.getColumnLetterFromString("Sample time [ms]")
            self.tolerance_col_letter = self.getColumnLetterFromString("Tolerance [%]")
        except ValueError:
            print("In file ", self.tc_file_name)
            exit()


def importFunctionFiles(fileName, sheetName):
    wbFunctions = load_workbook(fileName)
    wsFunctions = wbFunctions[sheetName]
    checkActionFile(worksheetAction=wsFunctions, worksheetActionFileName=fileName)
    ## migliorare la gestione degli errori con python - try and catch

    print(wbFunctions.sheetnames)
    return wsFunctions


def importBuildFile(fileName, sheetName):
    wbBuild = load_workbook(fileName)
    wsBuild = wbBuild[sheetName]
    checkBuildFile(worksheetBuild=wsBuild, worksheetBuildFileName=fileName)
    return wbBuild, wsBuild


def generateRunFileFromBuildFile(workbookBuild, sheetNameBuild, run_filename):
    workbookBuild.save(filename=run_filename)
    wbRun = load_workbook(run_filename)
    wsRun = wbRun[sheetNameBuild]
    return wbRun, wsRun
