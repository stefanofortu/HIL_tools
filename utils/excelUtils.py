from openpyxl.utils import get_column_letter


def getColumnLetterFromString(worksheet, columnName):
    for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for colNum, cell in enumerate(row, start=1):
            if cell == columnName:
                return get_column_letter(colNum)
    print("Column " + columnName + " not found")
    raise ValueError


def getColumnIndexFromString(worksheet, columnName):
    for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for colNum, cell in enumerate(row, start=1):
            if cell == columnName:
                return colNum
    print("Column " + columnName + " not found")
    raise ValueError


def remove_empty_consecutive_rows(inString):
    """ in_rows_list : input cell array """
    in_rows_list = inString.split('\n')

    """ print input string row by row """
    # for i, elem in enumerate(in_rows_list):
    #     print(str(i) + " : " + elem)
    # print("len :" + str(len(in_rows_list)))

    """ fill a list with the indexes of the not-empty rows """
    list_not_empty_rows_index = [i for i, elem in enumerate(in_rows_list) if len(elem) != 0]
    # print(list_not_empty_rows_index)

    """ output cell array : each element is a row """
    out_rows_list = []

    for i in range(0, len(list_not_empty_rows_index)):

        """ fill the output cell array with only the not-empty rows of the input array """
        out_rows_list.append(in_rows_list[list_not_empty_rows_index[i]])

        """ if two consecutive not-empty rows has distances greater than 2, that means that there were
        at least one empty rows between them: we need to re-introduce (only one!) between the rows
        So:
        if before at the last element of the array """
        if i < len(list_not_empty_rows_index) - 1:
            # if the distance among the indexes is greater than 2, add one row
            if list_not_empty_rows_index[i + 1] >= list_not_empty_rows_index[i] + 2:
                out_rows_list.append("")

    # print output string row by row
    # for i, elem in enumerate(out_rows_list):
    #     print(str(i) + " : " + elem)
    # print("len :" + str(len(out_rows_list)))

    """ join all the rows in a string """
    output_string = '\n'.join(out_rows_list)

    return output_string


def fix_bullet_lists(inString):
    """ in_rows_list : input cell array """
    in_rows_list = inString.split('\n')

    # """ print input string row by row """
    # for i, elem in enumerate(in_rows_list):
    #     print(str(i) + " : " + elem)
    # print("len :" + str(len(in_rows_list)))

    """ output cell array : each element is a row """
    out_rows_list = []

    for r in in_rows_list:
        """ find the position of the first 'point' """
        pos_of_first_point = r.find('.')
        if pos_of_first_point != -1:
            """ create a string containing only the number """
            string_with_number = r[0:pos_of_first_point]
            """ create a string with all the rest of the string, point excluded"""
            rest_of_the_row = r[pos_of_first_point + 1:]
            """ if first of the point there is a number and after the point there is not an empty row"""
            if string_with_number.isnumeric() and rest_of_the_row.strip() != "":
                """ add a row with only number + "." """
                out_rows_list.append(string_with_number + ".")
                """ add the rest of the row as a new row"""
                out_rows_list.append(rest_of_the_row.strip())
            else:
                out_rows_list.append(r)
        else:
            out_rows_list.append(r)

    # """print output string row by row """
    # for i, elem in enumerate(out_rows_list):
    #     print(str(i) + " : " + elem)
    # print("len :" + str(len(out_rows_list)))

    """ join all the rows in a string """
    output_string = '\n'.join(out_rows_list)

    return output_string


def polarion_to_excel_conversion(inString):
    """ remove polarion character \r """
    output_string = inString.replace("\r", "")
    """ remove excel character : _x000d_ """
    output_string = output_string.replace("_x000d_ ", "")
    return output_string


def excel_to_polarion_conversion(inString):
    """ rimuove i \r di Polarion """
    output_string = inString.replace("\n", "<br>")
    return output_string
