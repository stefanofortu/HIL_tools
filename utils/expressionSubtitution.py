from openpyxl.styles import Alignment

from utils.fileTemplateConfiguration import file_TC_BUILD_Column
from utils.excelUtils import getColumnIndexFromString, getColumnLetterFromString
from utils.importDictionary import singleTestStep
from copy import copy
from openpyxl.utils import get_column_letter


def findExpressions(wsStart, substitutionDictionary):
    columnPreconditionIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['precondition_header'])
    columnExpectedResIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['expected_header'])
    for row in wsStart.iter_rows(values_only=True,
                                 min_col=columnPreconditionIndex,
                                 max_col=columnExpectedResIndex):  #
        for cell in row:
            if isinstance(cell, str) and len(cell) > 2:
                if "()" in cell:
                    if cell not in substitutionDictionary:
                        print("function " + cell + " not found in dictionary")


def substituteFunctions(wsStart, wsEnd, substitutionDictionary, copyStyle=False):
    columnEnableIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['enable_header'])
    columnTestNIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['testN_header'])
    columnTestIDIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['testID_header'])
    columnDescriptionIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['stepDescr_header'])
    columnPreconditionIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['precondition_header'])
    columnActionIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['action_header'])
    columnExpectedResIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['expected_header'])
    columnTimeStepIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['timeStep_header'])
    columnSampleTimeIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['sampleTime_header'])
    columnToleranceIndex = getColumnIndexFromString(wsStart, file_TC_BUILD_Column['tolerance_header'])

    columnEnableLetter = getColumnLetterFromString(wsStart, file_TC_BUILD_Column['enable_header'])
    columnTestNLetter = getColumnLetterFromString(wsStart, file_TC_BUILD_Column['testN_header'])
    columnDescriptionLetter = getColumnLetterFromString(wsStart, file_TC_BUILD_Column['stepDescr_header'])
    columnPreconditionLetter = getColumnLetterFromString(wsStart, file_TC_BUILD_Column['precondition_header'])
    columnActionLetter = getColumnLetterFromString(wsStart, file_TC_BUILD_Column['action_header'])
    columnExpectedResLetter = getColumnLetterFromString(wsStart, file_TC_BUILD_Column['expected_header'])
    columnTimeStepLetter = getColumnLetterFromString(wsStart, file_TC_BUILD_Column['timeStep_header'])
    columnSampleTimeLetter = getColumnLetterFromString(wsStart, file_TC_BUILD_Column['sampleTime_header'])
    columnToleranceLetter = getColumnLetterFromString(wsStart, file_TC_BUILD_Column['tolerance_header'])

    # startColIndex = column_index_from_string(startCol)
    # endColIndex = column_index_from_string(endCol)
    c = 0
    translation = wsStart.max_row + 15
    rowNumEnd = translation
    rowsAdded = 0
    for rowNumStart, row in enumerate(wsStart.iter_rows(min_col=columnPreconditionIndex,
                                                        max_col=columnExpectedResIndex), start=1):
        foundStringToSubstitute = False
        for colNum, cell in enumerate(row, start=columnPreconditionIndex):
            if isinstance(cell.value, str):
                # print(wsheetEnd.cell(row=rowNumEnd, column=colNum + startColIndex).value)
                if len(cell.value) > 2:
                    if "()" in cell.value:
                        functionParsed = parseFunction(cell.value)
                        functionName = functionParsed['functionName']
                        parametersList = functionParsed['parametersList']
                        substitutedFunction = createFunction(functionParsed, substitutionDictionary)
                        if functionName in substitutionDictionary:
                            foundStringToSubstitute = True
                            # enableCell = worksheetStart.cell(row=rowNumStart, column=columnEnableIndex)
                            testNCell = wsStart.cell(row=rowNumStart, column=columnTestNIndex)
                            # testIDCell = worksheetStart.cell(row=rowNumStart, column=columnTestIDIndex)
                            # parametersList
                            # for dataNum, t in enumerate(substitutionDictionary[functionName]['data']):
                            for dataNum, t in enumerate(substitutedFunction):
                                # riporto il valore di "step" in
                                newRowPos = rowNumStart + translation + rowsAdded + dataNum

                                # riporto il valore di "enable "
                                wsEnd.cell(row=newRowPos, column=columnEnableIndex, value=t.enable)
                                if cell.has_style and copyStyle:
                                    wsEnd[columnEnableLetter + str(newRowPos)].alignment = Alignment(
                                        horizontal='center')

                                # print(enableCell.value)
                                # riporto il valore in testN
                                wsEnd.cell(row=newRowPos, column=columnTestNIndex, value=testNCell.value)
                                if copyStyle:
                                    wsEnd[columnTestNLetter + str(newRowPos)].alignment = copy(testNCell.alignment)
                                # Alignment(horizontal='center')
                                ## riporto il valore in testID
                                # worksheetEnd.cell(row=newRowPos, column=columnTestIDIndex, value=testIDCell.value)

                                # riporto il valore della descrizione nella colonna della step description
                                wsEnd.cell(row=newRowPos, column=columnDescriptionIndex, value=t.descr)
                                if cell.has_style and copyStyle:
                                    wsEnd[columnDescriptionLetter + str(newRowPos)].fill = copy(cell.fill)
                                    wsEnd[columnDescriptionLetter + str(newRowPos)].border = copy(cell.border)
                                    wsEnd[columnDescriptionLetter + str(newRowPos)].font = copy(cell.font)
                                # riporto il valore dell'azione/risultato atteso" nella colonna corrente
                                wsEnd.cell(row=newRowPos, column=colNum, value=t.act)
                                if cell.has_style and copyStyle:
                                    wsEnd[columnPreconditionLetter + str(newRowPos)].fill = copy(cell.fill)
                                    wsEnd[columnPreconditionLetter + str(newRowPos)].border = copy(cell.border)
                                    wsEnd[columnPreconditionLetter + str(newRowPos)].font = copy(cell.font)
                                    wsEnd[columnActionLetter + str(newRowPos)].fill = copy(cell.fill)
                                    wsEnd[columnActionLetter + str(newRowPos)].border = copy(cell.border)
                                    wsEnd[columnActionLetter + str(newRowPos)].font = copy(cell.font)
                                # riporto il valore di "expected res" nella colonna degli expected results
                                wsEnd.cell(row=newRowPos, column=columnExpectedResIndex, value=t.exp)
                                if cell.has_style and copyStyle:
                                    wsEnd[columnExpectedResLetter + str(newRowPos)].fill = copy(cell.fill)
                                    wsEnd[columnExpectedResLetter + str(newRowPos)].border = copy(cell.border)
                                    wsEnd[columnExpectedResLetter + str(newRowPos)].font = copy(cell.font)

                                # riporto il valore di "time step"
                                wsEnd.cell(row=newRowPos, column=columnTimeStepIndex, value=t.timeStep)
                                if cell.has_style and copyStyle:
                                    wsEnd[columnTimeStepLetter + str(newRowPos)].fill = copy(cell.fill)
                                    wsEnd[columnTimeStepLetter + str(newRowPos)].border = copy(cell.border)
                                    wsEnd[columnTimeStepLetter + str(newRowPos)].font = copy(cell.font)
                                    wsEnd[columnTimeStepLetter + str(newRowPos)].alignment = Alignment(
                                        horizontal='center')

                                # riporto il valore di "sample time"
                                wsEnd.cell(row=newRowPos, column=columnSampleTimeIndex, value=t.sampleTime)
                                if cell.has_style and copyStyle:
                                    wsEnd[columnSampleTimeLetter + str(newRowPos)].fill = copy(cell.fill)
                                    wsEnd[columnSampleTimeLetter + str(newRowPos)].border = copy(cell.border)
                                    wsEnd[columnSampleTimeLetter + str(newRowPos)].font = copy(cell.font)
                                    wsEnd[columnSampleTimeLetter + str(newRowPos)].alignment = Alignment(
                                        horizontal='center')

                                # riporto il valore di tolerance nella colonna degli expected results
                                wsEnd.cell(row=newRowPos, column=columnToleranceIndex, value=t.tolerance)
                                if cell.has_style and copyStyle:
                                    wsEnd[columnToleranceLetter + str(newRowPos)].fill = copy(cell.fill)
                                    wsEnd[columnToleranceLetter + str(newRowPos)].border = copy(cell.border)
                                    wsEnd[columnToleranceLetter + str(newRowPos)].font = copy(cell.font)
                                    wsEnd[columnToleranceLetter + str(newRowPos)].alignment = Alignment(
                                        horizontal='center')

                            rowsAdded += len(substitutedFunction) - 1
        if not foundStringToSubstitute:
            wsEnd.move_range("A" + str(rowNumStart) + ":AB" + str(rowNumStart),
                             rows=translation + rowsAdded, cols=0)

    wsEnd.delete_rows(1, amount=translation)


def parseFunction(stringToParse):
    if "=" in stringToParse:
        [functionName, parameters] = stringToParse.split("=")
        if ";" in parameters:
            parametersList = parameters.split(";")
        else:
            parametersList = [parameters]
    else:
        functionName = stringToParse
        parametersList = []

    return {"functionName": functionName, 'parametersList': parametersList}


def createFunction(functionParsed, substitutionDictionary):
    functionParsedName = functionParsed['functionName']
    parametersParsedList = functionParsed['parametersList']
    substitutedFunction = []
    if functionParsedName in substitutionDictionary:
        if len(substitutionDictionary[functionParsedName]['parameters']) == len(parametersParsedList):
            for functionStep in substitutionDictionary[functionParsedName]['data']:
                descr = functionStep.descr
                act = functionStep.act
                exp = functionStep.exp
                for parIndex, par in enumerate(substitutionDictionary[functionParsedName]['parameters']):
                    if isinstance(descr, str):
                        descr = descr.replace("{" + par + "}", parametersParsedList[parIndex])
                    if isinstance(act, str):
                        act = act.replace("{" + par + "}", parametersParsedList[parIndex])
                    if isinstance(exp, str):
                        exp = exp.replace("{" + par + "}", parametersParsedList[parIndex])
                s = singleTestStep(enable=functionStep.enable, descr=descr, act=act, exp=exp,
                                   timeStep=functionStep.timeStep,
                                   sampleTime=functionStep.sampleTime, tolerance=functionStep.tolerance)
                substitutedFunction.append(s)
        else:
            print(functionParsedName)
            raise ValueError
    else:
        print(functionParsedName + " not found")
        raise KeyError
    return substitutedFunction


def disableSequences(worksheet):
    columnEnableIndex = getColumnIndexFromString(worksheet, file_TC_BUILD_Column['enable_header'])
    columnTcTypeIndex = getColumnIndexFromString(worksheet, file_TC_BUILD_Column['testType_header'])

    columnTestNindxex = getColumnIndexFromString(worksheet, file_TC_BUILD_Column['testN_header'])

    # copia gli ID e il test number nella colonna successiva a quella della riga "Manual"
    testEnabled = ""
    for col in worksheet.iter_cols(min_col=columnTcTypeIndex, max_col=columnTcTypeIndex):
        for rowNum, cell in enumerate(col, start=1):
            if cell.value == "Manual":
                testEnabled = worksheet.cell(row=rowNum, column=columnEnableIndex).value
            else:
                if testEnabled == "OFF":
                    worksheet.cell(row=rowNum, column=columnEnableIndex, value="OFF")


def removeTestTypeColumn(worksheet):
    columnTcTypeIndex = getColumnIndexFromString(worksheet, file_TC_BUILD_Column['testType_header'])
    columnTestIDIndex = getColumnIndexFromString(worksheet, file_TC_BUILD_Column['testID_header'])
    columnTestNindxex = getColumnIndexFromString(worksheet, file_TC_BUILD_Column['testN_header'])

    # copia gli ID e il test number nella colonna successiva a quella della riga "Manual"
    testIdToJoin = []
    for col in worksheet.iter_cols(min_col=columnTcTypeIndex, max_col=columnTcTypeIndex):
        for rowNum, cell in enumerate(col, start=1):
            if cell.value == "Manual":
                testID = worksheet.cell(row=rowNum, column=columnTestIDIndex).value
                testIdToJoin.append(testID)
                testN = worksheet.cell(row=rowNum, column=columnTestNindxex).value
            else:
                if len(testIdToJoin) > 0:
                    testIDs = ""
                    if len(testIdToJoin) == 1:
                        testIDs = testIdToJoin[0]
                    else:
                        for id in testIdToJoin:
                            if id is not None:
                                testIDs = testIDs + id + ";"
                            else:
                                print(id, rowNum, columnTestIDIndex)
                                exit()
                    worksheet.cell(row=rowNum, column=columnTestIDIndex, value=testIDs)
                    worksheet.cell(row=rowNum, column=columnTestNindxex, value=testN)
                    testIdToJoin.clear()

                # worksheet.cell(row=rowNum, column=columnTestIDIndex, value=testID)
                # testN = worksheet.cell(row=rowNum, column=columnTestNindxex).value
                # worksheet.cell(row=rowNum, column=columnTestNindxex, value=testN)
                # # enable = worksheet.cell(row=currentRow, column=1).value
                # # worksheet.cell(row=currentRow + 1, column=1, value=enable)
                # rowToDelete = []
                # testIdToJoin = []

    # rimuove le righe "Manual"
    currentRow = 1
    for col in worksheet.iter_cols(min_col=columnTcTypeIndex, max_col=columnTcTypeIndex):
        for cell in col:
            if cell.value == "Manual":
                worksheet.delete_rows(currentRow)
            else:
                currentRow += 1

    # rimuove la colonna "TC Type e riformatta il file di conseguenza"
    worksheet.delete_cols(columnTcTypeIndex)
    columnCurrent = columnTcTypeIndex
    while columnCurrent <= worksheet.max_column:
        i = get_column_letter(columnCurrent)
        j = get_column_letter(columnCurrent + 1)
        worksheet.column_dimensions[i].width = copy(worksheet.column_dimensions[j].width)
        columnCurrent += 1
