from openpyxl import Workbook
from importDictionary import importDictionary
from importDictionaryV2 import importDictionaryV2
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from expressionSubtitution import substituteExpressionsByRows, findExpressions, substituteExpressionsByRowsV2
from utils.verifyTemplates import checkBuildFile, checkRunFile, checkActionFile

function_action_filename = "C:\\Users\\Stefano\\Desktop\\WIP\\HIL_1805\\HIL\\Documentazione\\3.Tools\\F175_actions_2021_05_15.xlsx"
function_action_sheet = "actions"

function_verify_filename = "C:\\Users\\Stefano\\Desktop\\WIP\\HIL_1805\\HIL\\Documentazione\\3.Tools\\F175_AF_verify_2021_05_15.xlsx"
function_verify_sheet1 = "verify"
function_verify_sheet2 = "AF"
function_verify_sheet3 = "verify_doors"

function_verify_New_filename = "C:\\Users\\Stefano\\Desktop\\WIP\\HIL_1805\\HIL\\Documentazione\\3.Tools\\NewProposal_Verify.xlsx"
function_verify_New_sheetName = "verify"
# source_filename = "C:\\Users\\Stefano\\Desktop\\WIP\\HIL_1805\\HIL\\TC_AF\\2-TC_Build\\TC_AF_HIL_Build_2021_05_13.xlsx"
# source_sheet = "TC_HIL_EU"
# destination_filename = "C:\\Users\\Stefano\\Desktop\\WIP\\HIL_1805\\HIL\\TC_AF\\3-TC_Run\\TC_AF_HIL_Run_2021_05_13.xlsx"


build_filename = "C:\\Users\\Stefano\\Desktop\\WIP\\HIL_1805\\HIL\\TC_AF\\2-TC_Build\\TC_AF_CheckConf_Build.xlsx"
source_sheet = "TC_AF_Configuration"
run_filename = "C:\\Users\\Stefano\\Desktop\\WIP\\HIL_1805\\HIL\\TC_AF\\3-TC_Run\\TC_AF_CheckConf_Run.xlsx"

# " Carica i file delle funzioni, foglio per foglio"
functionDictionary = {}

wb2 = load_workbook(function_action_filename)
ws = wb2[function_action_sheet]

res = importDictionary(worksheet=ws, functionDictionary=functionDictionary, dictionaryType="action")
if res == 1:
    print("MINOR: Error found in " + function_action_filename + ",sheet : " + function_action_sheet)

wb2 = load_workbook(function_verify_filename)
ws = wb2[function_verify_sheet1]
res = importDictionary(worksheet=ws, functionDictionary=functionDictionary, dictionaryType="verify")
if res == 1:
    print("MINOR: Error found in " + function_verify_filename + ",sheet : " + function_verify_sheet1)

ws = wb2[function_verify_sheet2]
res = importDictionary(worksheet=ws, functionDictionary=functionDictionary, dictionaryType="verify")
if res == 1:
    print("MINOR: Error found in " + function_verify_filename + ",sheet : " + function_verify_sheet2)

ws = wb2[function_verify_sheet3]
res = importDictionary(worksheet=ws, functionDictionary=functionDictionary, dictionaryType="verify")
if res == 1:
    print("MINOR: Error found in " + function_verify_filename + ",sheet : " + function_verify_sheet3)

functionDictionary = {}
wbActionNew = load_workbook(function_verify_New_filename)
wsActionNew = wbActionNew[function_verify_New_sheetName]
checkActionFile(worksheetAction=wsActionNew, worksheetActionFileName=function_verify_New_filename)

res = importDictionaryV2(worksheetAction=wsActionNew, functionDictionary=functionDictionary, dictionaryType="verify")
if res == 1:
    print("MINOR: Error found in " + function_verify_filename + ",sheet : " + function_verify_sheet3)


wbBuild = load_workbook(build_filename)
wbBuild.save(filename=run_filename)
wbRun = load_workbook(run_filename)

wsBuild = wbBuild[source_sheet]
wsRun = wbRun[source_sheet]

checkBuildFile(worksheetBuild=wsBuild, worksheetBuildFileName=build_filename)

findExpressions(worksheetStart=wsBuild,
                substitutionDictionary=functionDictionary)

print("dizionario acquisito")
substituteExpressionsByRowsV2(worksheetStart=wsBuild,
                              worksheetEnd=wsRun,
                              substitutionDictionary=functionDictionary)
print("sostituzione fatta")
# " Salva"
wbRun.save(filename=run_filename)
