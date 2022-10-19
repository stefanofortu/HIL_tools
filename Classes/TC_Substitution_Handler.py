from Classes.Configuration_Data import TC_Substitution_Configuration_Data
from utils.excelUtils import getColumnIndexFromString, remove_empty_consecutive_rows, fix_bullet_lists
from openpyxl import load_workbook
from utils.fileTemplateConfiguration import file_TC_MANUAL_Column


class TC_Substitution_Handler:
    def __init__(self, configuration_data=None):
        if isinstance(configuration_data, TC_Substitution_Configuration_Data):
            self.cfg_data = configuration_data
        else:
            self.cfg_data = TC_Substitution_Configuration_Data()

    @staticmethod
    def replace_cell(stringList, find_array, replace_array):
        outCell = []
        workCell = []
        for elem in stringList:
            workCell.append(elem)

        row_num = 0
        while row_num < len(workCell):
            if row_num <= (len(workCell) - len(find_array)):
                match_counter = 0
                for find_array_index in range(0, len(find_array)):
                    if workCell[row_num + find_array_index] == find_array[find_array_index]:
                        match_counter += 1
                if match_counter == len(find_array):
                    # print("====")
                    # print(workCell[row_num-1])
                    # print(workCell[row_num])
                    # print(workCell[row_num+1])
                    # print("====")
                    # print(len(find_array))
                    for replace_array_index in range(0, len(replace_array)):
                        outCell.append(replace_array[replace_array_index])
                    row_num += len(find_array)
                else:
                    outCell.append(workCell[row_num])
                    row_num += 1
            else:
                outCell.append(workCell[row_num])
                row_num += 1

        workCell = []
        for elem in outCell:
            workCell.append(elem)

        return outCell

    def exec_substitution(self):
        """ This is a quick summary line used as a description of the object.
        quick summary line used as a description of the object
        quick summary line used as a description of the object
        quick summary line used as a description of the object
        """

        wb_in = load_workbook(self.cfg_data.input_file_path)
        ws_in = wb_in[self.cfg_data.input_file_sheet]

        print("import input file : DONE")

        wb_find_replace = load_workbook(self.cfg_data.find_replace_file_path)
        ws_find_replace = wb_find_replace[self.cfg_data.find_replace_file_sheet]

        print("open find replace file : DONE")
        substitution_list_from_excel = []
        for row in ws_find_replace.iter_rows(min_row=2, min_col=1, max_col=2):
            single_substitution = {"find": "string", "replace": "string"}
            for colNum, cell in enumerate(row):
                if isinstance(cell.value, str):
                    if colNum == 0:
                        single_substitution["find"] = cell.value.split("\n")
                    elif colNum == 1:
                        single_substitution["replace"] = cell.value.split("\n")
            substitution_list_from_excel.append(single_substitution)

        print(substitution_list_from_excel)
        print("import find replace file : DONE")

        wb_in.save(filename=self.cfg_data.output_file_path)
        wbOut = load_workbook(self.cfg_data.output_file_path)
        wsOut = wbOut[self.cfg_data.input_file_sheet]

        print("saved Copy of Input file: DONE")

        columnPreconditionIndex = getColumnIndexFromString(wsOut, file_TC_MANUAL_Column['precondition_header'])
        columnActionIndex = getColumnIndexFromString(wsOut, file_TC_MANUAL_Column['action_header'])
        columnExpectedResIndex = getColumnIndexFromString(wsOut, file_TC_MANUAL_Column['expected_header'])

        for substitution in substitution_list_from_excel:
            find_array = substitution['find']
            replace_array = substitution['replace']

            for col in wsOut.iter_cols(min_row=1, min_col=columnPreconditionIndex,
                                       max_col=columnExpectedResIndex):
                for rowNum, cell in enumerate(col):
                    # tmp_array = []
                    # print(type(cellString))
                    if isinstance(cell.value, str):
                        cellString = cell.value
                        # print(len(cellString))
                        splitvalue = cellString.split('\n')
                        newCellValue = self.replace_cell(splitvalue, find_array, replace_array)
                        cellNewValueString = '\n'.join(newCellValue)
                        # print(len(cellNewValueString))
                        # print(cellNewValueString)
                        cell.value = str(cellNewValueString)
                        # print(cellNewValueString)
                        # tmp_array.append(row + "\n")
                        # print(tmp_array)
                        # Write in rich text
                        # optsheet.write_rich_string('A1', red, splitvalue[0], splitvalue[1])
                        # for elem in newCellValue:
                        # print("==================")
                        # optsheet.write_rich_string(rowNum + 1, 2, *tmp_array)
                    # Split characters
                    # print(cell.value)
            # optbook.close()

        print("substitution : DONE")
        wbOut.save(filename=self.cfg_data.output_file_path)

        print("file output saving : DONE")

    def exec_cleanup(self):

        wb_in = load_workbook(self.cfg_data.input_file_path)
        ws_in = wb_in[self.cfg_data.input_file_sheet]

        print("import input file : DONE")
        wb_in.save(filename=self.cfg_data.output_file_path)

        wbOut = load_workbook(self.cfg_data.output_file_path)
        wsOut = wbOut[self.cfg_data.input_file_sheet]

        print("saved Copy of Input file: DONE")

        columnPreconditionIndex = getColumnIndexFromString(wsOut, file_TC_MANUAL_Column['precondition_header'])
        columnExpectedResIndex = getColumnIndexFromString(wsOut, file_TC_MANUAL_Column['expected_header'])

        for col in wsOut.iter_cols(min_row=1, max_row=1640,
                                   min_col=columnPreconditionIndex,
                                   max_col=columnExpectedResIndex):
            for cell in col:
                if isinstance(cell.value, str):
                    cellString = cell.value
                    cellNewValueString = remove_empty_consecutive_rows(cellString)
                    cell.value = str(cellNewValueString)

        print("cancellazioni righe vuote : DONE")

        wbOut.save(filename=self.cfg_data.output_file_path)

        print("file output saving : DONE")

    def exec_bullet_lists_fix(self):

        wb_in = load_workbook(self.cfg_data.input_file_path)
        ws_in = wb_in[self.cfg_data.input_file_sheet]

        print("import input file : DONE")
        wb_in.save(filename=self.cfg_data.output_file_path)

        wbOut = load_workbook(self.cfg_data.output_file_path)
        wsOut = wbOut[self.cfg_data.input_file_sheet]

        print("saved Copy of Input file: DONE")

        columnPreconditionIndex = getColumnIndexFromString(wsOut, file_TC_MANUAL_Column['precondition_header'])
        columnExpectedResIndex = getColumnIndexFromString(wsOut, file_TC_MANUAL_Column['expected_header'])

        for col in wsOut.iter_cols(min_row=1, max_row=1640,
                                   min_col=columnPreconditionIndex,
                                   max_col=columnExpectedResIndex):
            for cell in col:
                if isinstance(cell.value, str):
                    cellString = cell.value
                    cellNewValueString = fix_bullet_lists(cellString)
                    cell.value = str(cellNewValueString)

        print("sistemazione elenchi puntati : DONE")

        wbOut.save(filename=self.cfg_data.output_file_path)

        print("file output saving : DONE")
