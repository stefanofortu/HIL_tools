from openpyxl.styles import Alignment

from utils.fileTemplateConfiguration import file_TC_RUN_Column
from utils.excelUtils import getColumnIndexFromString


def fillStepIDCounter(worksheet):
    columnStepIDIndex = getColumnIndexFromString(worksheet, file_TC_RUN_Column['stepID_header'])

    for col in worksheet.iter_cols(min_col=columnStepIDIndex, max_col=columnStepIDIndex):
        for rowNum, cell in enumerate(col):
            if rowNum == 0:
                continue
            cell.value = rowNum
            cell.alignment = Alignment(horizontal='center')


#############################################
#
# Se la colonna Enable è vuota, scrivi "ON"
#
##############################################
def fillEnableColumn(worksheet):
    columnEnableIndex = getColumnIndexFromString(worksheet, file_TC_RUN_Column['enable_header'])

    for col in worksheet.iter_cols(min_col=columnEnableIndex, max_col=columnEnableIndex):
        for rowNum, cell in enumerate(col):
            if rowNum == 0:
                continue
            if cell.value is None or cell.value == "":
                cell.value = "ON"
                cell.alignment = Alignment(horizontal='center')


#############################################
#
# Se la colonna TEST N è vuota,
# scrive il valore della riga sopra
#
##############################################
def fillTestNColumn(worksheet):
    columnTestNIndex = getColumnIndexFromString(worksheet, file_TC_RUN_Column['testN_header'])
    previousCellValue = None

    for col in worksheet.iter_cols(min_col=columnTestNIndex, max_col=columnTestNIndex):
        for rowNum, cell in enumerate(col):
            if rowNum == 0:
                previousCellValue = cell.value
                continue
            if cell.value is None or cell.value == "":
                cell.value = previousCellValue
                cell.alignment = Alignment(horizontal='center')
            previousCellValue = cell.value

